"""
Genera DATABASE_RELATIONS.xlsx con 3 hojas:
  1. Relaciones      — lista completa de FK filtrable
  2. Resumen nodos   — cuántas FK entran/salen de cada tabla
  3. Matriz          — tabla-origen vs tabla-destino (cuenta de FKs)
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from collections import defaultdict

# ── Datos extraídos de la BD ──────────────────────────────────────────────────
RELATIONS = [
    # accounting
    ("accounting","account_balances","account_code","accounting","chart_of_accounts","account_code"),
    ("accounting","chart_of_accounts","parent_account_code","accounting","chart_of_accounts","account_code"),
    ("accounting","credit_installments","operation_id","accounting","credit_operations","id"),
    ("accounting","credit_installments","settlement_id","accounting","credit_settlements","id"),
    ("accounting","credit_operations","journal_entry_id","accounting","journal_entries","id"),
    ("accounting","credit_operations","provider_id","accounting","external_credit_providers","id"),
    ("accounting","credit_settlements","journal_entry_id","accounting","journal_entries","id"),
    ("accounting","credit_settlements","operation_id","accounting","credit_operations","id"),
    ("accounting","journal_entry_lines","account_code","accounting","chart_of_accounts","account_code"),
    ("accounting","journal_entry_lines","journal_entry_id","accounting","journal_entries","id"),
    ("accounting","monthly_reports","period_closure_id","accounting","period_closures","id"),
    ("accounting","partner_compensation_runs","journal_entry_id","accounting","journal_entries","id"),
    ("accounting","payroll_payments","journal_entry_id","accounting","journal_entries","id"),
    ("accounting","payroll_payments","payroll_run_id","accounting","payroll_runs","id"),
    ("accounting","payroll_runs","journal_entry_id","accounting","journal_entries","id"),
    ("accounting","tax_config","account_code_credit","accounting","chart_of_accounts","account_code"),
    ("accounting","tax_config","account_code_debit","accounting","chart_of_accounts","account_code"),
    # catalog
    ("catalog","_mig_category_map","catalog_id","catalog","categories","id"),
    ("catalog","_mig_pack_map","catalog_product_id","catalog","products","id"),
    ("catalog","_mig_product_map","catalog_id","catalog","products","id"),
    ("catalog","_mig_subcategory_map","catalog_id","catalog","categories","id"),
    ("catalog","categories","parent_id","catalog","categories","id"),
    ("catalog","external_catalog_sync_runs","source_id","catalog","external_catalog_sources","id"),
    ("catalog","product_bundles","bundle_product_id","catalog","products","id"),
    ("catalog","product_bundles","component_product_id","catalog","products","id"),
    ("catalog","product_documents","product_id","catalog","products","id"),
    ("catalog","products","category_id","catalog","categories","id"),
    ("catalog","products","tax_rate_id","catalog","tax_rates","id"),
    ("catalog","stock_alerts","product_id","catalog","products","id"),
    ("catalog","stock_movements","product_id","catalog","products","id"),
    # crm
    ("crm","client_notes","client_id","crm","clients","id"),
    ("crm","contacts","client_id","crm","clients","id"),
    ("crm","interactions","client_id","crm","clients","id"),
    ("crm","location_notes","location_id","crm","location","id"),
    # internal
    ("internal","authorized_users","linked_employee_id","internal","employees","id"),
    ("internal","authorized_users","linked_partner_id","internal","partners","id"),
    ("internal","invitation_tokens","user_id","internal","authorized_users","id"),
    ("internal","partner_compensation_runs","created_by","internal","authorized_users","id"),
    ("internal","partner_compensation_runs","partner_id","internal","partners","id"),
    ("internal","partner_payroll_profiles","partner_id","internal","partners","id"),
    ("internal","partner_payroll_profiles","updated_by","internal","authorized_users","id"),
    ("internal","payroll_settings","updated_by","internal","authorized_users","id"),
    ("internal","payroll_settings_audit","changed_by","internal","authorized_users","id"),
    ("internal","payroll_settings_audit","settings_id","internal","payroll_settings","id"),
    ("internal","product_pack_items","pack_id","internal","product_packs","id"),
    ("internal","product_pack_items","product_id","internal","products","id"),
    ("internal","product_subcategories","category_id","internal","product_categories","id"),
    ("internal","products","category_id","internal","product_categories","id"),
    ("internal","products","default_tax_id","internal","taxes","id"),
    ("internal","products","subcategory_id","internal","product_subcategories","id"),
    ("internal","report_settings","updated_by","internal","authorized_users","id"),
    ("internal","task_activity","task_id","internal","tasks","id"),
    ("internal","task_assignees","task_id","internal","tasks","id"),
    ("internal","technicians","created_by","internal","authorized_users","id"),
    ("internal","user_roles","granted_by","internal","authorized_users","id"),
    ("internal","user_roles","role_id","internal","roles","id"),
    ("internal","user_roles","user_id","internal","authorized_users","id"),
    # projects
    ("projects","av_projects","project_id","projects","projects","id"),
    ("projects","customer_orders","project_id","projects","projects","id"),
    ("projects","expenses","project_id","projects","projects","id"),
    ("projects","project_activity","project_id","projects","projects","id"),
    ("projects","project_documents","project_id","projects","projects","id"),
    ("projects","project_sites","project_id","projects","projects","id"),
    ("projects","project_tasks","parent_task_id","projects","project_tasks","id"),
    ("projects","project_tasks","project_id","projects","projects","id"),
    ("projects","projects","default_site_id","projects","project_sites","id"),
    ("projects","site_technician_assignments","site_id","projects","project_sites","id"),
    ("projects","site_visits","site_id","projects","project_sites","id"),
    # quotes
    ("quotes","quote_activity","quote_id","quotes","quotes","id"),
    ("quotes","quote_history","quote_id","quotes","quotes","id"),
    ("quotes","quote_lines","quote_id","quotes","quotes","id"),
    # sales
    ("sales","invoice_lines","invoice_id","sales","invoices","id"),
    ("sales","invoice_payments","invoice_id","sales","invoices","id"),
    ("sales","invoices","original_invoice_id","sales","invoices","id"),
    ("sales","invoices","rectified_by_invoice_id","sales","invoices","id"),
    ("sales","purchase_invoice_lines","purchase_invoice_id","sales","purchase_invoices","id"),
    ("sales","purchase_invoice_payments","purchase_invoice_id","sales","purchase_invoices","id"),
    ("sales","purchase_order_lines","purchase_order_id","sales","purchase_orders","id"),
    ("sales","purchase_orders","linked_purchase_invoice_id","sales","purchase_invoices","id"),
]

# ── Colores por esquema ───────────────────────────────────────────────────────
SCHEMA_COLORS = {
    "accounting": "D6E4FF",
    "audit":      "FFF3CD",
    "catalog":    "D4EDDA",
    "crm":        "FCE4EC",
    "internal":   "E8D5F5",
    "projects":   "D1ECF1",
    "public":     "F5F5F5",
    "quotes":     "FFF0D6",
    "sales":      "FFE0E0",
    "security":   "E2E2E2",
}
SCHEMA_HEADER_COLORS = {
    "accounting": "4472C4",
    "audit":      "F4B942",
    "catalog":    "2E7D32",
    "crm":        "C62828",
    "internal":   "7B1FA2",
    "projects":   "00838F",
    "public":     "616161",
    "quotes":     "E65100",
    "sales":      "B71C1C",
    "security":   "424242",
}

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(color="FFFFFF"):
    return Font(bold=True, color=color, size=10)

def make_workbook():
    wb = openpyxl.Workbook()

    # ── HOJA 1: Relaciones ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Relaciones"

    headers = [
        "Esquema origen", "Tabla origen", "Columna FK",
        "Esquema destino", "Tabla destino", "Columna destino",
        "Tipo relación"
    ]
    col_widths = [16, 32, 28, 16, 32, 20, 16]

    # Título
    ws1.merge_cells("A1:G1")
    title_cell = ws1["A1"]
    title_cell.value = "NEXO AV — Mapa de Relaciones FK entre tablas"
    title_cell.font = Font(bold=True, size=13, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="D6DCE4")
    ws1.row_dimensions[1].height = 22

    # Subtítulo
    ws1.merge_cells("A2:G2")
    sub = ws1["A2"]
    sub.value = f"Total FK: {len(RELATIONS)}  |  Fuente: Supabase information_schema  |  Fecha: 2026-03-26"
    sub.font = Font(italic=True, size=9, color="595959")
    sub.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 14

    # Cabecera
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws1.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        c.border = thin_border()
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[3].height = 18

    # Datos
    for row_i, (fs, ft, fc, ts, tt, tc) in enumerate(RELATIONS, 4):
        # tipo de relación
        if ft == tt and fs == ts:
            rel_type = "Auto-ref"
        elif tc == "id":
            rel_type = "N → 1"
        else:
            rel_type = "N → N"

        values = [fs, ft, fc, ts, tt, tc, rel_type]
        row_color = SCHEMA_COLORS.get(fs, "FFFFFF")
        for col, val in enumerate(values, 1):
            c = ws1.cell(row=row_i, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=row_color)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center")
            c.font = Font(size=9)
        ws1.row_dimensions[row_i].height = 14

    # Autofilter
    ws1.auto_filter.ref = f"A3:G{3 + len(RELATIONS)}"
    ws1.freeze_panes = "A4"

    # ── HOJA 2: Resumen nodos ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen nodos")

    # Calcular métricas por tabla
    fk_out = defaultdict(int)   # FKs que salen (from_table)
    fk_in  = defaultdict(int)   # FKs que entran (to_table)
    schemas = {}

    for fs, ft, fc, ts, tt, tc in RELATIONS:
        key_from = f"{fs}.{ft}"
        key_to   = f"{ts}.{tt}"
        fk_out[key_from] += 1
        fk_in[key_to]    += 1
        schemas[key_from] = fs
        schemas[key_to]   = ts

    all_tables = sorted(set(list(fk_out.keys()) + list(fk_in.keys())))

    ws2.merge_cells("A1:F1")
    t2 = ws2["A1"]
    t2.value = "Resumen de conectividad por tabla"
    t2.font = Font(bold=True, size=13, color="1F3864")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    t2.fill = PatternFill("solid", fgColor="D6DCE4")
    ws2.row_dimensions[1].height = 22

    h2 = ["Esquema", "Tabla", "FK salientes (referencias a otras)", "FK entrantes (referenciada por)", "Total conexiones", "Rol principal"]
    cw2 = [16, 34, 30, 30, 18, 20]
    for col, (h, w) in enumerate(zip(h2, cw2), 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = thin_border()
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[2].height = 28

    for row_i, tbl in enumerate(all_tables, 3):
        schema = schemas.get(tbl, "")
        name   = tbl.split(".", 1)[1] if "." in tbl else tbl
        out_c  = fk_out.get(tbl, 0)
        in_c   = fk_in.get(tbl, 0)
        total  = out_c + in_c
        if in_c > out_c * 2:
            role = "Entidad central"
        elif out_c > in_c:
            role = "Tabla dependiente"
        elif out_c == 0 and in_c > 0:
            role = "Tabla raíz"
        else:
            role = "Intermedia"

        row_color = SCHEMA_COLORS.get(schema, "FFFFFF")
        for col, val in enumerate([schema, name, out_c, in_c, total, role], 1):
            c = ws2.cell(row=row_i, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=row_color)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", horizontal="center" if col > 2 else "left")
            c.font = Font(size=9)
        ws2.row_dimensions[row_i].height = 14

    ws2.auto_filter.ref = f"A2:F{1 + len(all_tables)}"
    ws2.freeze_panes = "A3"

    # ── HOJA 3: Matriz de adyacencia ──────────────────────────────────────────
    ws3 = wb.create_sheet("Matriz")

    # Solo tablas con relaciones, abreviadas como schema.table
    matrix_tables = sorted(set(
        [f"{fs}.{ft}" for fs, ft, *_ in RELATIONS] +
        [f"{ts}.{tt}" for _, _, _, ts, tt, _ in RELATIONS]
    ))

    # Construir mapa
    rel_map = defaultdict(int)
    for fs, ft, fc, ts, tt, tc in RELATIONS:
        rel_map[(f"{fs}.{ft}", f"{ts}.{tt}")] += 1

    ws3.merge_cells(f"A1:{get_column_letter(len(matrix_tables)+1)}1")
    t3 = ws3["A1"]
    t3.value = "Matriz de adyacencia FK  (fila = ORIGEN, columna = DESTINO, número = cantidad de FK)"
    t3.font = Font(bold=True, size=11, color="1F3864")
    t3.alignment = Alignment(horizontal="center")
    t3.fill = PatternFill("solid", fgColor="D6DCE4")
    ws3.row_dimensions[1].height = 20

    # Cabeceras columna (destinos)
    ws3.cell(row=2, column=1, value="ORIGEN \\ DESTINO").font = Font(bold=True, size=8)
    ws3.column_dimensions["A"].width = 34

    for col_i, tbl in enumerate(matrix_tables, 2):
        schema = tbl.split(".")[0]
        c = ws3.cell(row=2, column=col_i, value=tbl)
        c.font = Font(bold=True, size=7, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=SCHEMA_HEADER_COLORS.get(schema, "1F3864"))
        c.alignment = Alignment(textRotation=90, horizontal="center", vertical="bottom")
        ws3.column_dimensions[get_column_letter(col_i)].width = 6
    ws3.row_dimensions[2].height = 90

    # Filas (orígenes)
    for row_i, from_tbl in enumerate(matrix_tables, 3):
        schema = from_tbl.split(".")[0]
        lc = ws3.cell(row=row_i, column=1, value=from_tbl)
        lc.font = Font(bold=True, size=8, color="FFFFFF")
        lc.fill = PatternFill("solid", fgColor=SCHEMA_HEADER_COLORS.get(schema, "1F3864"))
        lc.alignment = Alignment(horizontal="left", vertical="center")

        for col_i, to_tbl in enumerate(matrix_tables, 2):
            count = rel_map.get((from_tbl, to_tbl), 0)
            c = ws3.cell(row=row_i, column=col_i, value=count if count > 0 else "")
            if count > 0:
                intensity = min(int(255 - count * 40), 230)
                hex_green = f"{intensity:02X}FF{intensity:02X}" if count < 4 else "00AA00"
                c.fill = PatternFill("solid", fgColor=hex_green)
                c.font = Font(bold=True, size=9, color="1F3864")
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = PatternFill("solid", fgColor="F8F8F8")
            c.border = thin_border()
        ws3.row_dimensions[row_i].height = 14

    ws3.freeze_panes = "B3"

    # ── Guardar ───────────────────────────────────────────────────────────────
    output = r"C:\Users\AlexBurgues\AV TECH ESDEVENIMENTS SL\Marketing - Documentos\V2_WEB\av-tech-p-gina-de-inicio\docs\architecture\DATABASE_RELATIONS.xlsx"
    wb.save(output)
    print(f"Guardado: {output}")
    print(f"  Hoja 1 'Relaciones':    {len(RELATIONS)} filas")
    print(f"  Hoja 2 'Resumen nodos': {len(all_tables)} tablas")
    print(f"  Hoja 3 'Matriz':        {len(matrix_tables)}×{len(matrix_tables)}")

if __name__ == "__main__":
    make_workbook()
